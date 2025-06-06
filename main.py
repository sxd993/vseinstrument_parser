from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from urllib.parse import urljoin, parse_qs, urlencode, urlparse

def setup_driver():
    """Настройка Selenium-драйвера."""
    options = Options()
    options.add_argument('--headless')  # Фоновый режим
    options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36')
    options.add_argument('--disable-blink-features=AutomationControlled')  # Обход обнаружения автоматизации
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    return driver

def get_page_content(driver, url):
    """Загрузка страницы и возврат HTML."""
    driver.get(url)
    time.sleep(5)  # Увеличено время ожидания для динамического контента
    return BeautifulSoup(driver.page_source, 'html.parser')

def parse_products(soup, max_products, current_product_count):
    """Извлечение данных о товарах из HTML с учетом лимита."""
    products_data = []
    # Попробуем несколько селекторов для товаров
    products = soup.find_all('div', attrs={'data-qa': 'products-tile'}) or \
               soup.find_all('div', class_='product-card') or \
               soup.find_all('div', class_='product-tile')
    
    if not products:
        return products_data, current_product_count

    for product in products:
        if max_products > 0 and current_product_count >= max_products:
            break
        try:
            # Извлечение данных с резервными селекторами
            code = product.find('p', attrs={'data-qa': 'product-code-text'}).text.strip() if product.find('p', attrs={'data-qa': 'product-code-text'}) else \
                   product.find('span', class_='product-code').text.strip() if product.find('span', class_='product-code') else 'N/A'
            name = product.find('a', attrs={'data-qa': 'product-name'}).text.strip() if product.find('a', attrs={'data-qa': 'product-name'}) else \
                   product.find('a', class_='product-name').text.strip() if product.find('a', class_='product-name') else 'N/A'
            url = product.find('a', attrs={'data-qa': 'product-name'})['href'] if product.find('a', attrs={'data-qa': 'product-name'}) else \
                  product.find('a', class_='product-name')['href'] if product.find('a', class_='product-name') else 'N/A'
            price = product.find('p', attrs={'data-qa': 'product-price-current'}).text.strip() if product.find('p', attrs={'data-qa': 'product-price-current'}) else \
                    product.find('span', class_='price').text.strip() if product.find('span', class_='price') else 'N/A'
            rating = product.find('input', attrs={'name': 'rating'})['value'] if product.find('input', attrs={'name': 'rating'}) else \
                     product.find('span', class_='rating').text.strip() if product.find('span', class_='rating') else 'N/A'
            brand = product.find('span', class_='brand').text.strip() if product.find('span', class_='brand') else 'N/A'

            # Если URL относительный, преобразуем в абсолютный
            if url != 'N/A' and not url.startswith('http'):
                url = urljoin('https://www.vseinstrumenti.ru', url)

            # Добавляем данные о товаре
            products_data.append({
                "Артикул": code,
                "Название": name,
                "Бренд": brand,
                "URL": url,
                "Обычная цена": price,
                "Рейтинг": rating
            })

            current_product_count += 1

        except AttributeError:
            continue
    
    return products_data, current_product_count

def get_next_page(soup, base_url):
    """Поиск URL следующей страницы с сохранением параметров запроса."""
    try:
        # Парсим базовый URL для извлечения параметров
        parsed_url = urlparse(base_url)
        query_params = parse_qs(parsed_url.query)
        
        # Проверяем различные варианты пагинации
        pagination = soup.find('div', class_='pagination') or \
                     soup.find('div', class_='pagination-wrapper') or \
                     soup.find('nav', class_='pagination')
        if not pagination:
            return None
            
        next_page = pagination.find('a', class_='next-page') or \
                    pagination.find('a', attrs={'data-qa': 'next-page'}) or \
                    pagination.find('a', class_='pagination__next')
        if next_page and 'href' in next_page.attrs:
            next_url = next_page['href']
            # Если URL относительный, преобразуем в абсолютный
            if not next_url.startswith('http'):
                next_url = urljoin(base_url, next_url)
            
            # Парсим URL следующей страницы
            parsed_next_url = urlparse(next_url)
            next_query_params = parse_qs(parsed_next_url.query)
            
            # Объединяем параметры из исходного URL и следующей страницы
            combined_params = query_params.copy()
            combined_params.update(next_query_params)
            
            # Формируем новый URL с сохранением параметров
            new_query = urlencode(combined_params, doseq=True)
            next_url = f"{parsed_next_url.scheme}://{parsed_next_url.netloc}{parsed_next_url.path}?{new_query}"
            
            return next_url
        return None
    except Exception:
        return None

def save_to_excel(products_data, filename='products.xlsx'):
    """Сохранение данных в Excel."""
    try:
        # Если данных нет, создаем пустой DataFrame с заголовками
        if not products_data:
            products_data = [{
                "Артикул": "N/A",
                "Название": "Нет данных",
                "Бренд": "N/A",
                "URL": "N/A",
                "Обычная цена": "N/A",
                "Рейтинг": "N/A"
            }]
        
        # Преобразуем данные о товарах в DataFrame
        df = pd.DataFrame(products_data)

        # Сохраняем в Excel
        with pd.ExcelWriter(filename, engine="openpyxl", mode="w") as writer:
            df.to_excel(writer, sheet_name="Товары", index=False)

        # Настройка ширины столбцов
        workbook = load_workbook(filename)
        worksheet = workbook["Товары"]
        default_width = 8.43
        first_width = float(default_width * 1.3)
        third_width = float(default_width * 6)

        columns_to_widen_one = ["Артикул", "Бренд", "Обычная цена", "Рейтинг"]
        columns_to_widen_three = ["Название", "URL"]
        for col_idx, col_name in enumerate(df.columns, start=1):
            column_letter = get_column_letter(col_idx)
            if col_name in columns_to_widen_one:
                worksheet.column_dimensions[column_letter].width = first_width
            elif col_name in columns_to_widen_three:
                worksheet.column_dimensions[column_letter].width = third_width

        # Форматирование заголовков
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        workbook.save(filename)

    except Exception:
        raise

def main(url, max_products=0):
    """Основная функция парсинга с лимитом товаров."""
    driver = setup_driver()
    products_data = []
    current_product_count = 0
    base_url = url.split('#')[0]  # Убираем хэш-параметры

    try:
        current_url = base_url
        while current_url and (max_products == 0 or current_product_count < max_products):
            soup = get_page_content(driver, current_url)
            if not soup:
                break

            # Парсинг товаров
            page_products, current_product_count = parse_products(soup, max_products, current_product_count)
            products_data.extend(page_products)

            # Проверка лимита товаров
            if max_products > 0 and current_product_count >= max_products:
                break

            # Поиск следующей страницы
            current_url = get_next_page(soup, base_url)
            time.sleep(3)  # Задержка для этичного парсинга

    finally:
        driver.quit()

    # Сохранение в Excel даже при отсутствии данных
    save_to_excel(products_data)

if __name__ == "__main__":
    url = input("Введите URL для парсинга (например, https://www.vseinstrumenti.ru/category/perforatory-32/): ")
    max_products = int(input("Введите максимальное количество товаров для парсинга (0 для всех): "))
    main(url, max_products)