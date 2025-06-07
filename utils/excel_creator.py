"""
Модуль для сохранения спарсенных данных о товарах в Excel-файл.
Этот модуль:
- Преобразует данные о товарах в pandas DataFrame.
- Форматирует Excel-файл с настроенными ширинами столбцов и стилизованными заголовками.
- Обрабатывает ошибки при создании Excel и логирует результаты.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from utils.logger import logger


def save_to_excel(products_data, filename="products.xlsx"):
    """
    Сохранение данных о товарах в Excel-файл.

    Аргументы:
        products_data (list): Список словарей с данными о товарах.
        filename (str): Имя выходного Excel-файла.

    Исключения:
        Exception: Если создание Excel-файла не удалось.
    """
    try:
        df = pd.DataFrame(products_data)
        with pd.ExcelWriter(filename, engine="openpyxl", mode="w") as writer:
            df.to_excel(writer, sheet_name="Товары", index=False)

        workbook = load_workbook(filename)
        worksheet = workbook["Товары"]

        # Настройка ширины столбцов и стилизация заголовков
        for col_idx, col_name in enumerate(df.columns, start=1):
            column_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[column_letter].width = 20
            worksheet[f"{column_letter}1"].font = Font(bold=True)
            worksheet[f"{column_letter}1"].alignment = Alignment(horizontal="center")

        workbook.save(filename)
        logger.info(f"Excel-файл сохранен: {filename}")
    except Exception as e:
        logger.error(f"Ошибка сохранения Excel: {str(e)}")
        raise
